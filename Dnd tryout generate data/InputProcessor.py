# InputProcessor.py
"""
Evaluation-only InputProcessor

What it does:
- Reads your dataset Excel (Parsing Examples sheet)
- For each row:
    1) Precheck classification
    2) If precheck == clear: intent parsing
    3) Canonicalize outputs, resolve UIDs->names, score with BLEU + ROUGE-L
- Writes/updates an output workbook with per-row metrics + a Summary sheet:
    - Means + Standard Deviations for metrics (BLEU/ROUGE, time, cost, tokens)
    - Welch t-tests on BLEU + ROUGE between precheck/single/multi groups

Key features:
- Resume support: skips Example # already present in the output file
- Handles trailing “used but empty” Excel rows (stops after N blank Example # rows)
- Summary-only mode: rebuild Summary (and optionally clean junk rows) with ZERO model calls

Important requested change:
- BLEU is adaptive:
    * truth_type == "precheck" -> BLEU-1
    * otherwise -> BLEU-4
"""

from __future__ import annotations

import math
import random
import re
import time
from pathlib import Path
from typing import Dict, Iterable, List, Tuple, Optional, Any

import openpyxl
import openai
import httpx  # OpenRouter /generation fallback

import gameSetup
import gameRenderer
from config import base_url, key, model_parsing, model_precheck, showPrints, useAI


# =========================
# USER CONFIG
# =========================

# If True: run exactly 1 TRUTH precheck + 1 TRUTH single + 1 TRUTH multi
TEST_MODE: bool = False
TEST_SEED: int = 42

# Dataset
INPUT_XLSX_PATH: str = r"C:\Users\arthu\OneDrive\Desktop\Classes\Master\Module 5+6\Human-Machine dialogue\Dnd tryout generate data\parsing ground truth.xlsx"
SHEET_NAME: str = "Parsing Examples"

# Output naming
OUTPUT_SUFFIX: str = "_eval_results.xlsx"

# OpenRouter usage accounting (tries to include usage.cost + usage tokens in response)
ENABLE_USAGE_ACCOUNTING: bool = True

# Save after each row (safe resume)
SAVE_EVERY_ROW: bool = True

# Stop dataset iteration after N consecutive blank Example # rows
TRAILING_EMPTY_BREAK: int = 10

# Summary-only mode (no model calls; just clean + recompute Summary)
RECALC_SUMMARY_ONLY: bool = True
CLEAN_INVALID_ROWS: bool = True


# =========================
# INTERNAL GLOBALS
# =========================

_EVAL_ALREADY_RUN: bool = False

_UID_MAPS_BUILT = False
_CHAR_UID_TO_NAME: Dict[str, str] = {}
_ITEM_UID_TO_NAME: Dict[str, str] = {}
_AREA_UID_TO_NAME: Dict[str, str] = {}

EXPECTED_FIELDS: Tuple[str, ...] = (
    "action",
    "requested_action",
    "target",
    "indirect_target",
    "item",
    "location",
    "topic_of_conversation",
)

_DATASET_LABELS = {"clear", "long", "insufficient", "impossible", "question", "undo"}

OUTPUT_COLUMNS = [
    "Example #",
    "Input",
    "Truth Output",
    "Hypothesis Output",
    "Truth Type",
    "Hypothesis Type",
    "BLEU",  # adaptive: BLEU-1 for precheck truth, BLEU-4 otherwise
    "ROUGE-L",
    "Precheck Time (s)",
    "Parsing Time (s)",
    "Total Time (s)",
    "Precheck Cost (credits)",
    "Parsing Cost (credits)",
    "Total Cost (credits)",
    "Precheck Prompt Tokens",
    "Precheck Completion Tokens",
    "Precheck Total Tokens",
    "Parsing Prompt Tokens",
    "Parsing Completion Tokens",
    "Parsing Total Tokens",
    "Total Prompt Tokens",
    "Total Completion Tokens",
    "Total Tokens",
]

NUMERIC_COLUMNS = [
    "BLEU",
    "ROUGE-L",
    "Precheck Time (s)",
    "Parsing Time (s)",
    "Total Time (s)",
    "Precheck Cost (credits)",
    "Parsing Cost (credits)",
    "Total Cost (credits)",
    "Precheck Prompt Tokens",
    "Precheck Completion Tokens",
    "Precheck Total Tokens",
    "Parsing Prompt Tokens",
    "Parsing Completion Tokens",
    "Parsing Total Tokens",
    "Total Prompt Tokens",
    "Total Completion Tokens",
    "Total Tokens",
]


# =========================
# SMALL HELPERS
# =========================

def _is_blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


# =========================
# WORLD HELPERS
# =========================

def _iter_all_areas() -> Iterable[object]:
    for v in vars(gameSetup).values():
        try:
            if isinstance(v, gameRenderer.SubArea):
                yield v
        except Exception:
            continue


def _build_uid_maps() -> None:
    global _UID_MAPS_BUILT, _CHAR_UID_TO_NAME, _ITEM_UID_TO_NAME, _AREA_UID_TO_NAME
    if _UID_MAPS_BUILT:
        return

    char_map: Dict[str, str] = {}
    item_map: Dict[str, str] = {}
    area_map: Dict[str, str] = {}

    for a in _iter_all_areas():
        auid = getattr(a, "uid", None)
        aname = getattr(a, "name", None)
        if isinstance(auid, str) and auid and isinstance(aname, str) and aname:
            area_map[auid] = aname

        for it in getattr(a, "key_items", []) or []:
            iuid = getattr(it, "uid", None)
            iname = getattr(it, "name", None)
            if isinstance(iuid, str) and iuid and isinstance(iname, str) and iname:
                item_map[iuid] = iname

        for c in getattr(a, "characters", []) or []:
            cuid = getattr(c, "uid", None)
            cname = getattr(c, "name", None)
            if isinstance(cuid, str) and cuid and isinstance(cname, str) and cname:
                char_map[cuid] = cname

            for it in getattr(c, "inventory", []) or []:
                iuid = getattr(it, "uid", None)
                iname = getattr(it, "name", None)
                if isinstance(iuid, str) and iuid and isinstance(iname, str) and iname:
                    item_map[iuid] = iname

    _CHAR_UID_TO_NAME = char_map
    _ITEM_UID_TO_NAME = item_map
    _AREA_UID_TO_NAME = area_map
    _UID_MAPS_BUILT = True


def _strip_wrapping_quotes(s: str) -> str:
    s2 = str(s).strip()
    if len(s2) >= 2 and ((s2[0] == s2[-1] == '"') or (s2[0] == s2[-1] == "'")):
        return s2[1:-1].strip()
    return s2


def _resolve_uid_to_name(value: str) -> str:
    if value is None:
        return "0"
    v = _strip_wrapping_quotes(value)
    if v in ("", "0", "None", "null"):
        return "0"

    _build_uid_maps()

    if v in _CHAR_UID_TO_NAME:
        return _CHAR_UID_TO_NAME[v]
    if v in _ITEM_UID_TO_NAME:
        return _ITEM_UID_TO_NAME[v]
    if v in _AREA_UID_TO_NAME:
        return _AREA_UID_TO_NAME[v]

    # conservative embedded replacement
    for uid_map in (_CHAR_UID_TO_NAME, _ITEM_UID_TO_NAME, _AREA_UID_TO_NAME):
        for uid, name in uid_map.items():
            if uid and uid in v:
                return v.replace(uid, name)

    return v


def _resolve_fields_to_names(fields: Dict[str, str]) -> Dict[str, str]:
    out = dict(fields)
    for k in ("target", "indirect_target", "item", "location"):
        if k in out:
            out[k] = _resolve_uid_to_name(out[k])
    return out


def _resolve_uids_in_free_text(text: str) -> str:
    if not text:
        return text
    _build_uid_maps()
    out = str(text)
    merged = {}
    merged.update(_CHAR_UID_TO_NAME)
    merged.update(_ITEM_UID_TO_NAME)
    merged.update(_AREA_UID_TO_NAME)
    for uid, name in merged.items():
        if uid and uid in out:
            out = out.replace(uid, name)
    return out


# =========================
# NORMALIZATION
# =========================

def normalize_space(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def normalize_precheck_label(raw: str) -> str:
    s = normalize_space(raw).lower()
    s = re.split(r"\s+", s)[0] if s else ""
    s = s.strip(".,:;!\"'()[]{}")

    if s == "redo":
        return "undo"
    if s in ("unrelated", "irrelevant"):
        return "impossible"
    if s == "unknown":
        return "insufficient"

    if s in _DATASET_LABELS:
        return s

    if "question" in s:
        return "question"
    if "insuff" in s:
        return "insufficient"
    if "imposs" in s:
        return "impossible"
    if "long" in s:
        return "long"
    if "undo" in s:
        return "undo"
    return "clear"


# =========================
# COST + TOKENS (OpenRouter)
# =========================

def _openrouter_generation_cost(gen_id: str) -> Optional[float]:
    """Fallback: fetch total_cost via /generation?id=... (may not always be immediately ready)."""
    if not gen_id or not isinstance(gen_id, str):
        return None

    gen_url = base_url.rstrip("/") + "/generation"
    headers = {"Authorization": f"Bearer {key}"}
    params = {"id": gen_id}

    backoffs = [0.2, 0.4, 0.8, 1.2]
    with httpx.Client(timeout=20.0) as client:
        for wait_s in backoffs:
            try:
                r = client.get(gen_url, headers=headers, params=params)
                if r.status_code == 200:
                    data = (r.json() or {}).get("data") or {}
                    tc = data.get("total_cost")
                    if tc is not None:
                        try:
                            return float(tc)
                        except Exception:
                            return None
                    return None

                if r.status_code in (404, 429, 500, 502, 503):
                    time.sleep(wait_s)
                    continue

                return None
            except Exception:
                time.sleep(wait_s)
                continue

    return None


def _extract_cost(resp) -> float:
    usage = getattr(resp, "usage", None)
    if usage is not None:
        c = getattr(usage, "cost", None)
        if c is None and hasattr(usage, "model_dump"):
            try:
                c = usage.model_dump().get("cost")
            except Exception:
                c = None
        if c is not None:
            try:
                return float(c)
            except Exception:
                pass

    gen_id = getattr(resp, "id", None)
    if isinstance(gen_id, str) and gen_id:
        c2 = _openrouter_generation_cost(gen_id)
        if c2 is not None:
            return float(c2)

    return 0.0


def _extract_tokens(resp) -> Tuple[int, int, int]:
    """
    Returns (prompt_tokens, completion_tokens, total_tokens).
    If not present, returns (0,0,0).
    """
    usage = getattr(resp, "usage", None)
    if usage is None:
        return 0, 0, 0

    # usage can be an object with attrs, or something with model_dump
    def _get(u: Any, k: str) -> Optional[int]:
        val = getattr(u, k, None)
        if val is None and hasattr(u, "model_dump"):
            try:
                val = u.model_dump().get(k)
            except Exception:
                val = None
        try:
            return int(val) if val is not None else None
        except Exception:
            return None

    pt = _get(usage, "prompt_tokens") or 0
    ct = _get(usage, "completion_tokens") or 0
    tt = _get(usage, "total_tokens") or (pt + ct)
    return pt, ct, tt


# =========================
# OPENAI CALLS
# =========================

def _client() -> openai.OpenAI:
    return openai.OpenAI(api_key=key, base_url=base_url)


def AIprecheck_with_usage(user_text: str) -> Tuple[str, float, int, int, int]:
    """
    Returns:
      (label, cost, prompt_tokens, completion_tokens, total_tokens)
    """
    if not useAI:
        return "clear", 0.0, 0, 0, 0

    system = (
        "You are a safety/feasibility pre-check classifier for a text-adventure game.\n"
        "Return exactly ONE label (single word) from this set:\n"
        "Clear, Long, Insufficient, Impossible, Question, Undo\n\n"
        "Return ONLY the label."
    )

    client = _client()
    extra_body = {"usage": {"include": True}} if ENABLE_USAGE_ACCOUNTING else None

    last_resp = None
    for _ in range(3):
        resp = client.chat.completions.create(
            model=model_precheck,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user_text},
            ],
            temperature=0,
            max_tokens=16,
            extra_body=extra_body,
        )
        last_resp = resp
        txt = resp.choices[0].message.content or ""
        lab = normalize_precheck_label(txt)
        if lab in _DATASET_LABELS:
            cost = _extract_cost(resp)
            pt, ct, tt = _extract_tokens(resp)
            return lab, cost, pt, ct, tt

    # fallback
    if last_resp is not None:
        cost = _extract_cost(last_resp)
        pt, ct, tt = _extract_tokens(last_resp)
        return "clear", cost, pt, ct, tt

    return "clear", 0.0, 0, 0, 0


def _area_lines() -> str:
    lines: List[str] = []
    for a in _iter_all_areas():
        name = getattr(a, "name", "")
        if name:
            lines.append(f"Name: {name}")
    return "\n".join(lines)


def _char_lines() -> str:
    lines: List[str] = []
    for a in _iter_all_areas():
        aname = getattr(a, "name", "")
        for c in getattr(a, "characters", []) or []:
            cname = getattr(c, "name", "")
            if cname:
                lines.append(f"Name: {cname}, Area: {aname}")
    return "\n".join(lines)


def _item_lines() -> str:
    lines: List[str] = []
    for a in _iter_all_areas():
        aname = getattr(a, "name", "")
        for it in getattr(a, "key_items", []) or []:
            iname = getattr(it, "name", "")
            if iname:
                lines.append(f"Name: {iname}, Area: {aname}")
        for c in getattr(a, "characters", []) or []:
            cname = getattr(c, "name", "")
            for it in getattr(c, "inventory", []) or []:
                iname = getattr(it, "name", "")
                if iname:
                    lines.append(f"Name: {iname}, Holder: {cname}")
    return "\n".join(lines)


def AIparsing_with_usage(user_text: str) -> Tuple[str, float, int, int, int]:
    """
    Returns:
      (raw_output, cost, prompt_tokens, completion_tokens, total_tokens)
    """
    if not useAI:
        dummy = '1. "action:talk,requested_action:0,target:0,indirect_target:0,item:0,location:0,topic_of_conversation:0"'
        return dummy, 0.0, 0, 0, 0

    system = (
        "You are an intent parser for a text-adventure game.\n"
        "Output one or more intents in THIS EXACT FORMAT (match punctuation and ordering):\n"
        '1. "action:<...>,requested_action:<...>,target:<...>,indirect_target:<...>,item:<...>,location:<...>,topic_of_conversation:<...>"\n'
        "Rules:\n"
        "- Use 0 for any field that is not present.\n"
        "- Use NAMES (not IDs) for target/item/location.\n"
        "- Keep keys in the exact order shown.\n"
        "- Do not add extra commentary."
    )

    context = (
        "WORLD CONTEXT (names only):\n\n"
        f"AREAS:\n{_area_lines()}\n\n"
        f"CHARACTERS:\n{_char_lines()}\n\n"
        f"ITEMS:\n{_item_lines()}\n"
    )

    client = _client()
    extra_body = {"usage": {"include": True}} if ENABLE_USAGE_ACCOUNTING else None

    resp = client.chat.completions.create(
        model=model_parsing,
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": context + "\nUSER INPUT:\n" + user_text},
        ],
        temperature=0,
        max_tokens=300,
        extra_body=extra_body,
    )
    content = resp.choices[0].message.content or ""
    cost = _extract_cost(resp)
    pt, ct, tt = _extract_tokens(resp)
    return content, cost, pt, ct, tt


def timed_precheck(user_text: str) -> Tuple[str, float, float, int, int, int]:
    t0 = time.perf_counter()
    lab, cost, pt, ct, tt = AIprecheck_with_usage(user_text)
    t1 = time.perf_counter()
    return lab, float(t1 - t0), float(cost), pt, ct, tt


def timed_parsing(user_text: str) -> Tuple[str, float, float, int, int, int]:
    t0 = time.perf_counter()
    out, cost, pt, ct, tt = AIparsing_with_usage(user_text)
    t1 = time.perf_counter()
    return out, float(t1 - t0), float(cost), pt, ct, tt


# =========================
# CANONICALIZATION
# =========================

def _split_intent_blocks(text: str) -> List[str]:
    t = (text or "").strip()
    if not t:
        return []
    parts = re.split(r"(?=\b\d+\.)", t)
    blocks = [p.strip() for p in parts if p and p.strip()]
    if len(blocks) == 1 and not re.match(r"^\d+\.", blocks[0]):
        return [t]
    return blocks


def _robust_extract_fields(block: str) -> Dict[str, str]:
    b = re.sub(r"^\s*\d+\s*\.\s*", "", (block or "").strip())
    b = _strip_wrapping_quotes(b)
    b = b.replace("|", ",")

    fields: Dict[str, str] = {}
    for keyname in EXPECTED_FIELDS:
        m = re.search(rf"\b{re.escape(keyname)}\s*:\s*([^,\n\r\"]+)", b, flags=re.I)
        if m:
            fields[keyname] = _strip_wrapping_quotes(m.group(1).strip())

    for k in EXPECTED_FIELDS:
        fields.setdefault(k, "0")
    return fields


def canonicalize_parsing_output(raw: str) -> str:
    raw = raw or ""
    blocks = _split_intent_blocks(raw)
    if not blocks:
        return '1. "action:0,requested_action:0,target:0,indirect_target:0,item:0,location:0,topic_of_conversation:0"'

    out_lines: List[str] = []
    for i, b in enumerate(blocks, start=1):
        fields = _robust_extract_fields(b)
        fields = _resolve_fields_to_names(fields)

        for k in EXPECTED_FIELDS:
            fields[k] = normalize_space(fields.get(k, "0"))
            if fields[k] in ("", "None", "null"):
                fields[k] = "0"

        inner = ",".join([f"{k}:{fields[k]}" for k in EXPECTED_FIELDS])
        out_lines.append(f'{i}. "{inner}"')

    result = "\n".join(out_lines)
    return _resolve_uids_in_free_text(result)


def infer_intent_type(canonical: str) -> str:
    blocks = _split_intent_blocks(canonical)
    return "multi" if len(blocks) > 1 else "single"


# =========================
# SCORING (BLEU + ROUGE-L)
# =========================

def _tokenize(s: str) -> List[str]:
    s2 = normalize_space(s)
    if not s2:
        return []
    return re.findall(r"[A-Za-z0-9_]+|[^\w\s]", s2)


def _ngrams(tokens: List[str], n: int) -> List[Tuple[str, ...]]:
    return [tuple(tokens[i:i + n]) for i in range(len(tokens) - n + 1)]


def bleu(reference: str, hypothesis: str, max_n: int = 4, smooth: float = 1e-9) -> float:
    """
    BLEU up to max_n n-grams. Caps n to what is possible for short strings.
    """
    ref = _tokenize(reference)
    hyp = _tokenize(hypothesis)
    if not hyp or not ref:
        return 0.0

    max_n = int(max(1, min(max_n, len(hyp), len(ref))))

    precisions = []
    for n in range(1, max_n + 1):
        ref_ngrams = _ngrams(ref, n)
        hyp_ngrams = _ngrams(hyp, n)
        if not hyp_ngrams:
            precisions.append(smooth)
            continue

        ref_counts: Dict[Tuple[str, ...], int] = {}
        for ng in ref_ngrams:
            ref_counts[ng] = ref_counts.get(ng, 0) + 1

        hyp_counts: Dict[Tuple[str, ...], int] = {}
        for ng in hyp_ngrams:
            hyp_counts[ng] = hyp_counts.get(ng, 0) + 1

        match = 0
        for ng, c in hyp_counts.items():
            match += min(c, ref_counts.get(ng, 0))

        precisions.append((match + smooth) / (len(hyp_ngrams) + smooth))

    log_p = sum(math.log(p) for p in precisions) / float(max_n)
    geo_mean = math.exp(log_p)

    ref_len = len(ref)
    hyp_len = len(hyp)
    bp = 1.0 if hyp_len > ref_len else math.exp(1.0 - (ref_len / max(hyp_len, 1)))

    return float(bp * geo_mean)


def _lcs_length(a: List[str], b: List[str]) -> int:
    if not a or not b:
        return 0
    if len(a) < len(b):
        a, b = b, a
    prev = [0] * (len(b) + 1)
    for x in a:
        cur = [0]
        for j, y in enumerate(b, start=1):
            if x == y:
                cur.append(prev[j - 1] + 1)
            else:
                cur.append(max(prev[j], cur[-1]))
        prev = cur
    return prev[-1]


def rouge_l(reference: str, hypothesis: str) -> float:
    ref = _tokenize(reference)
    hyp = _tokenize(hypothesis)
    if not ref or not hyp:
        return 0.0
    lcs = _lcs_length(ref, hyp)
    prec = lcs / len(hyp)
    rec = lcs / len(ref)
    if prec + rec == 0:
        return 0.0
    beta = 1.2
    return ((1 + beta**2) * prec * rec) / (rec + beta**2 * prec)


# =========================
# OUTPUT / RESUME HELPERS
# =========================

def _get_output_path(input_path: str) -> Path:
    in_path = Path(str(input_path))
    return in_path.with_name(in_path.stem + OUTPUT_SUFFIX)


def _ensure_results_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    if "Eval Results" in wb.sheetnames:
        ws = wb["Eval Results"]
        if ws.max_row < 1:
            ws.append(OUTPUT_COLUMNS)
        return ws

    ws = wb.active
    ws.title = "Eval Results"
    ws.append(OUTPUT_COLUMNS)
    return ws


def _header_to_col_index(ws) -> Dict[str, int]:
    # find exact header row if possible
    for r in range(1, min(ws.max_row, 50) + 1):
        row = [ws.cell(r, c).value for c in range(1, len(OUTPUT_COLUMNS) + 1)]
        if row == OUTPUT_COLUMNS:
            return {name: i + 1 for i, name in enumerate(OUTPUT_COLUMNS)}
    return {name: i + 1 for i, name in enumerate(OUTPUT_COLUMNS)}


def _existing_example_ids(ws) -> set:
    col_map = _header_to_col_index(ws)
    ex_col = col_map.get("Example #", 1)
    existing = set()
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, ex_col).value
        if not _is_blank(val):
            existing.add(val)
    return existing


def _clean_eval_results_sheet(ws_out) -> int:
    """Delete rows where Example # is blank/None."""
    col_map = _header_to_col_index(ws_out)
    ex_col = col_map.get("Example #", 1)
    rows_to_delete = []
    for r in range(2, ws_out.max_row + 1):
        ex_id = ws_out.cell(r, ex_col).value
        if _is_blank(ex_id):
            rows_to_delete.append(r)
    for r in reversed(rows_to_delete):
        ws_out.delete_rows(r, 1)
    return len(rows_to_delete)


# =========================
# DATASET ITERATION (avoid trailing empty 'used' rows)
# =========================

def _iter_valid_dataset_rows(ws_in):
    """
    Iterate dataset rows but STOP after TRAILING_EMPTY_BREAK consecutive blank Example # rows.
    """
    empty_streak = 0
    for r in range(2, ws_in.max_row + 1):
        ex_id = ws_in.cell(r, 1).value
        if _is_blank(ex_id):
            empty_streak += 1
            if empty_streak >= TRAILING_EMPTY_BREAK:
                break
            continue
        empty_streak = 0
        yield r


# =========================
# TEST MODE SAMPLING
# =========================

def _truth_type_for_row(truth_pre: str, truth_parse: str) -> str:
    lab = normalize_precheck_label(str(truth_pre))
    if lab != "clear":
        return "precheck"
    canon = canonicalize_parsing_output(str(truth_parse))
    return infer_intent_type(canon)


def _pick_one_per_type(ws_in) -> List[int]:
    buckets: Dict[str, List[int]] = {"precheck": [], "single": [], "multi": []}

    for r in _iter_valid_dataset_rows(ws_in):
        truth_pre = ws_in.cell(r, 3).value or "Clear"
        truth_parse = ws_in.cell(r, 4).value or ""
        ttype = _truth_type_for_row(str(truth_pre), str(truth_parse))
        if ttype in buckets:
            buckets[ttype].append(r)

    rng = random.Random(TEST_SEED)
    picked: List[int] = []
    for t in ("precheck", "single", "multi"):
        if buckets[t]:
            picked.append(rng.choice(buckets[t]))

    # pad if missing types
    if len(picked) < 3:
        all_rows = list(_iter_valid_dataset_rows(ws_in))
        rng.shuffle(all_rows)
        for r in all_rows:
            if r not in picked:
                picked.append(r)
            if len(picked) >= 3:
                break

    return picked


# =========================
# SUMMARY + STATS + T-TESTS
# =========================

def _finite(xs: List[float]) -> List[float]:
    out = []
    for x in xs:
        try:
            xf = float(x)
            if not math.isnan(xf) and math.isfinite(xf):
                out.append(xf)
        except Exception:
            continue
    return out


def _mean(xs: List[float]) -> float:
    xs2 = _finite(xs)
    return float(sum(xs2) / len(xs2)) if xs2 else float("nan")


def _std(xs: List[float]) -> float:
    xs2 = _finite(xs)
    n = len(xs2)
    if n < 2:
        return float("nan")
    m = sum(xs2) / n
    var = sum((x - m) ** 2 for x in xs2) / (n - 1)
    return float(math.sqrt(var))


def _variance(xs: List[float]) -> float:
    xs2 = _finite(xs)
    n = len(xs2)
    if n < 2:
        return float("nan")
    m = sum(xs2) / n
    return float(sum((x - m) ** 2 for x in xs2) / (n - 1))


def _welch_ttest(a: List[float], b: List[float]) -> Tuple[float, float, float]:
    a = _finite(a)
    b = _finite(b)
    n1, n2 = len(a), len(b)
    if n1 < 2 or n2 < 2:
        return float("nan"), float("nan"), float("nan")

    m1, m2 = sum(a) / n1, sum(b) / n2
    v1, v2 = _variance(a), _variance(b)
    if not (v1 > 0 and v2 > 0):
        return float("nan"), float("nan"), float("nan")

    se = math.sqrt(v1 / n1 + v2 / n2)
    if se == 0:
        return float("nan"), float("nan"), float("nan")

    t = (m1 - m2) / se

    num = (v1 / n1 + v2 / n2) ** 2
    den = (v1 * v1) / (n1 * n1 * (n1 - 1)) + (v2 * v2) / (n2 * n2 * (n2 - 1))
    df = num / den if den != 0 else float("nan")

    # p-value via scipy if available else normal approx
    try:
        from scipy.stats import t as student_t  # type: ignore
        p = 2.0 * float(student_t.sf(abs(t), df))
        return float(t), float(df), float(p)
    except Exception:
        z = abs(t)
        p = 2.0 * (1.0 - 0.5 * (1.0 + math.erf(z / math.sqrt(2.0))))
        return float(t), float(df), float(p)


def _write_summary_sheet(wb_out: openpyxl.Workbook, results_ws) -> None:
    if "Summary" in wb_out.sheetnames:
        del wb_out["Summary"]
    ws = wb_out.create_sheet("Summary")

    col = _header_to_col_index(results_ws)

    rows = []
    for r in range(2, results_ws.max_row + 1):
        ex_id = results_ws.cell(r, col["Example #"]).value
        if _is_blank(ex_id):
            continue

        truth_type = results_ws.cell(r, col["Truth Type"]).value
        if _is_blank(truth_type):
            continue

        entry = {"Truth Type": str(truth_type)}
        for name in NUMERIC_COLUMNS:
            v = results_ws.cell(r, col[name]).value
            try:
                entry[name] = float(v)
            except Exception:
                entry[name] = float("nan")
        rows.append(entry)

    def group(filter_type: Optional[str]) -> List[dict]:
        if filter_type is None:
            return rows
        return [x for x in rows if x.get("Truth Type") == filter_type]

    groups = [
        ("overall", None),
        ("precheck", "precheck"),
        ("single", "single"),
        ("multi", "multi"),
    ]

    # Means + stds table
    header = ["group", "n"]
    for m in NUMERIC_COLUMNS:
        header.append(f"avg_{m}")
        header.append(f"std_{m}")
    ws.append(header)

    group_values: Dict[str, Dict[str, List[float]]] = {}

    for gname, gfilter in groups:
        g_rows = group(gfilter)
        n = len(g_rows)

        vals_map: Dict[str, List[float]] = {}
        row_out = [gname, n]
        for m in NUMERIC_COLUMNS:
            xs = [x[m] for x in g_rows]
            xs2 = _finite(xs)
            vals_map[m] = xs2
            row_out.append(_mean(xs2))
            row_out.append(_std(xs2))

        group_values[gname] = vals_map
        ws.append(row_out)

    ws.append([])
    ws.append(["Welch t-tests (two-sided) on BLEU and ROUGE-L"])
    ws.append(["metric", "group_a", "group_b", "t", "df", "p_value"])

    pairs = [("precheck", "single"), ("precheck", "multi"), ("single", "multi")]
    for metric in ("BLEU", "ROUGE-L"):
        for a, b in pairs:
            ta = group_values.get(a, {}).get(metric, [])
            tb = group_values.get(b, {}).get(metric, [])
            t_stat, df, p = _welch_ttest(ta, tb)
            ws.append([metric, a, b, t_stat, df, p])


# =========================
# EVALUATION
# =========================

def run_excel_evaluation(xlsx_path: str) -> str:
    out_path = _get_output_path(xlsx_path)

    # Summary-only mode (no new API calls)
    if RECALC_SUMMARY_ONLY and out_path.exists():
        wb_out = openpyxl.load_workbook(out_path)
        ws_out = wb_out["Eval Results"] if "Eval Results" in wb_out.sheetnames else wb_out.active

        removed = 0
        if CLEAN_INVALID_ROWS:
            removed = _clean_eval_results_sheet(ws_out)

        _write_summary_sheet(wb_out, ws_out)
        wb_out.save(out_path)

        if showPrints:
            print(f"[Summary-only] Rebuilt Summary. Removed {removed} invalid rows.")
        return str(out_path)

    # Normal evaluation mode
    wb_in = openpyxl.load_workbook(str(xlsx_path))
    if SHEET_NAME not in wb_in.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {wb_in.sheetnames}")
    ws_in = wb_in[SHEET_NAME]

    if out_path.exists():
        wb_out = openpyxl.load_workbook(out_path)
    else:
        wb_out = openpyxl.Workbook()

    ws_out = _ensure_results_sheet(wb_out)
    already = _existing_example_ids(ws_out)

    if TEST_MODE:
        target_rows = _pick_one_per_type(ws_in)
    else:
        target_rows = list(_iter_valid_dataset_rows(ws_in))

    rows_to_process: List[int] = []
    for r in target_rows:
        ex_id = ws_in.cell(r, 1).value
        if _is_blank(ex_id):
            continue
        if ex_id in already:
            continue
        rows_to_process.append(r)

    for r in rows_to_process:
        ex_id = ws_in.cell(r, 1).value
        player_inp = ws_in.cell(r, 2).value or ""
        truth_pre = ws_in.cell(r, 3).value or "Clear"
        truth_parse = ws_in.cell(r, 4).value or ""

        # safety skips
        if _is_blank(ex_id) or _is_blank(player_inp):
            continue

        # truth output + truth type
        truth_lab = normalize_precheck_label(str(truth_pre))
        if truth_lab != "clear":
            truth_out = truth_lab
            truth_type = "precheck"
        else:
            truth_out = canonicalize_parsing_output(str(truth_parse))
            truth_type = infer_intent_type(truth_out)

        # hypothesis precheck
        hypo_lab, t_pre, c_pre, pre_pt, pre_ct, pre_tt = timed_precheck(str(player_inp))

        # hypothesis parsing only if clear
        t_parse = 0.0
        c_parse = 0.0
        par_pt = par_ct = par_tt = 0

        if hypo_lab != "clear":
            hypo_out = hypo_lab
            hypo_type = "precheck"
        else:
            raw, t_parse, c_parse, par_pt, par_ct, par_tt = timed_parsing(str(player_inp))
            hypo_out = canonicalize_parsing_output(raw)
            hypo_type = infer_intent_type(hypo_out)

        total_t = float(t_pre + t_parse)
        total_c = float(c_pre + c_parse)

        total_pt = int(pre_pt + par_pt)
        total_ct = int(pre_ct + par_ct)
        total_tt = int(pre_tt + par_tt)

        # adaptive BLEU: BLEU-1 for truth precheck, else BLEU-4
        bleu_n = 1 if truth_type == "precheck" else 4
        b = bleu(truth_out, hypo_out, max_n=bleu_n)
        rL = rouge_l(truth_out, hypo_out)

        ws_out.append(
            [
                ex_id,
                str(player_inp),
                truth_out,
                hypo_out,
                truth_type,
                hypo_type,
                float(b),
                float(rL),
                float(t_pre),
                float(t_parse),
                float(total_t),
                float(c_pre),
                float(c_parse),
                float(total_c),
                int(pre_pt),
                int(pre_ct),
                int(pre_tt),
                int(par_pt),
                int(par_ct),
                int(par_tt),
                int(total_pt),
                int(total_ct),
                int(total_tt),
            ]
        )

        if showPrints:
            print(
                f"#{ex_id} truth_type={truth_type} hypo_type={hypo_type} "
                f"BLEU(n={bleu_n})={b:.3f} ROUGE-L={rL:.3f} "
                f"time={total_t:.3f}s cost={total_c:.6f} tokens={total_tt}"
            )

        if SAVE_EVERY_ROW:
            wb_out.save(out_path)

    if CLEAN_INVALID_ROWS:
        _clean_eval_results_sheet(ws_out)

    _write_summary_sheet(wb_out, ws_out)
    wb_out.save(out_path)
    return str(out_path)


# =========================
# MAIN ENTRY POINT (compat)
# =========================

def get_story(player_input: str):
    global _EVAL_ALREADY_RUN
    if _EVAL_ALREADY_RUN:
        return ("(Evaluation already ran.)", 1)

    _EVAL_ALREADY_RUN = True
    out_path = run_excel_evaluation(INPUT_XLSX_PATH)

    msg = (
        "✅ Done.\n"
        f"- Input:  {INPUT_XLSX_PATH}\n"
        f"- Output: {out_path}\n\n"
        "Notes:\n"
        "- Set RECALC_SUMMARY_ONLY=True to rebuild Summary instantly (no model calls).\n"
        "- BLEU is adaptive: precheck rows use BLEU-1; others use BLEU-4.\n"
        "- Summary now includes mean + standard deviation for all metrics (including time/cost/tokens)."
    )
    return (msg, 1)


if __name__ == "__main__":
    text, _ = get_story("")
    print(text)
